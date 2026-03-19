import asyncio
import random
from datetime import date, datetime, timedelta, time as dtime
from decimal import Decimal
from io import BytesIO

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook

from telegram import InlineKeyboardButton, InlineKeyboardMarkup, InputFile, Update
from telegram import ChatAction
from telegram.ext import (
    Application,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    filters,
)

from config import is_admin as is_admin_cfg, load_settings
from db import (
    add_trade,
    all_users,
    best_worst_day,
    daily_profit_series,
    active_users_for_motivation,
    ensure_admin,
    get_global_premium,
    init_db,
    last_trade_time,
    last_trades,
    list_users,
    set_user_ban,
    set_user_premium_days,
    set_user_show_on_leaderboard,
    set_user_leaderboard_paused,
    reset_leaderboard_all,
    set_global_premium,
    sum_profit,
    count_trades,
    count_users,
    count_trades_by_date,
    get_user_flags,
    get_user_record,
    leaderboard_totals,
    leaderboard_rank,
    inactive_users_for_reminder,
    trades_by_date,
    user_profit_summary_since,
    user_profit_summary,
    set_user_language,
    user_trades_range,
    clear_user_trades,
    extend_user_premium_days,
    pause_user_premium,
    resume_user_premium,
    cancel_user_premium,
    list_subscriptions,
    list_active_premium_users,
    upsert_user,
)
from i18n import SUPPORTED_LANGS, fallback_lang, get_first, pick as t
from logic import TradeInputs, compute_roi, compute_trade_profit, compute_qty, compute_weighted_averages
from ui import (
    back_cancel_keyboard,
    consent_leaderboard_keyboard,
    export_keyboard,
    main_menu_keyboard,
    days_range_keyboard,
    language_keyboard,
    result_keyboard,
    section_back_home_keyboard,
    trade_count_keyboard,
    trade_menu_keyboard,
    stats_menu_keyboard,
    tools_menu_keyboard,
    account_menu_keyboard,
    premium_dashboard_keyboard,
    premium_insights_keyboard,
    premium_reports_keyboard,
    escrow_result_keyboard,
    escrow_mode_keyboard,
)
from utils import clamp_range, fmt_money, parse_decimal


DEFAULT_AI_FALLBACK = "Good trade 🔥"


def translate_text(lang: str | None, key: str) -> str:
    """
    Safe wrapper around i18n `t(lang, key)` with a friendly fallback.
    """
    try:
        out = t(lang, key)
    except Exception:
        return DEFAULT_AI_FALLBACK
    if not out:
        return DEFAULT_AI_FALLBACK
    return out


def _format_pretty_date(dt: datetime | None) -> str:
    if dt is None:
        return "N/A"
    # Example: 20 Mar 2026
    return dt.strftime("%d %b %Y")


def _premium_labels(lang: str) -> dict[str, str]:
    lang = fallback_lang(lang)
    labels = {
        "en": {"title": "💎 Premium Activated", "start": "📅 Start", "exp": "⏳ Expires", "features": "✅ Features unlocked"},
        "hinglish": {"title": "💎 Premium Activated", "start": "📅 Start", "exp": "⏳ Expires", "features": "✅ Features unlocked"},
        "hi": {"title": "💎 Premium एक्टिवेट हो गया", "start": "📅 शुरू", "exp": "⏳ समाप्ति", "features": "✅ Features unlocked"},
        "zh": {"title": "💎 Premium 已激活", "start": "📅 开始", "exp": "⏳ 到期", "features": "✅ 已解锁功能"},
    }
    return labels.get(lang, labels["en"])


def _is_authorized_admin(user_id: int) -> bool:
    return is_admin_cfg(user_id)


def _days_left_from_expiry(expiry_raw: str | None) -> int | None:
    if not expiry_raw:
        return None
    try:
        expiry = datetime.fromisoformat(str(expiry_raw))
    except ValueError:
        return None
    delta = expiry - datetime.utcnow()
    return max(0, delta.days)


def _status_text(is_premium_flag: int, expiry_raw: str | None, is_paused_flag: int) -> str:
    if is_premium_flag != 1:
        return "Not Active"
    if is_paused_flag == 1:
        return "Paused"
    days_left = _days_left_from_expiry(expiry_raw)
    if days_left is None:
        return "Not Active"
    if days_left <= 0:
        return "Expired"
    return "Active"


async def db_to_thread(context: ContextTypes.DEFAULT_TYPE, fn, *args):
    return await asyncio.to_thread(fn, *args)


async def show_screen(update: Update, text: str, reply_markup: InlineKeyboardMarkup | None = None) -> None:
    """
    Single-screen UI helper:
    - For callback buttons: edit existing message.
    - For commands/normal text: send new message.
    """
    query = update.callback_query
    if query is not None and query.message is not None:
        try:
            await query.edit_message_text(text=text, reply_markup=reply_markup)
            return
        except Exception:
            # Fall back to sending a new message if edit fails.
            pass
    if update.effective_message is not None:
        await update.effective_message.reply_text(text, reply_markup=reply_markup)


async def show_session_screen(update: Update, context: ContextTypes.DEFAULT_TYPE, text: str, reply_markup: InlineKeyboardMarkup | None = None) -> None:
    """
    Edit a previously tracked UI message for text-driven flows (no spam).
    Falls back to regular show_screen when tracking is unavailable.
    """
    session = get_session(context)
    chat_id = session.get("ui_chat_id") if session else None
    message_id = session.get("ui_message_id") if session else None
    if chat_id and message_id:
        try:
            await context.bot.edit_message_text(chat_id=chat_id, message_id=message_id, text=text, reply_markup=reply_markup)
            return
        except Exception:
            pass
    await show_screen(update, text, reply_markup)


async def ensure_user(update: Update, context: ContextTypes.DEFAULT_TYPE, settings) -> None:
    user = update.effective_user
    if not user:
        return
    await db_to_thread(context, upsert_user, settings.db_path, user.id, user.username)
    flags = await db_to_thread(context, get_user_flags, settings.db_path, user.id)
    global_premium = await db_to_thread(context, get_global_premium, settings.db_path)
    context.user_data["lang"] = flags.get("language")
    context.user_data["is_banned"] = int(flags.get("is_banned") or 0)
    context.user_data["is_premium"] = int(flags.get("is_premium") or 0)
    context.user_data["global_premium"] = int(global_premium or 0)


def current_lang(context: ContextTypes.DEFAULT_TYPE) -> str:
    return fallback_lang(context.user_data.get("lang"))


def is_premium(context: ContextTypes.DEFAULT_TYPE) -> bool:
    return bool(
        int(context.user_data.get("is_premium") or 0)
        or int(context.user_data.get("global_premium") or 0)
    )


def days_limit(context: ContextTypes.DEFAULT_TYPE) -> int:
    # Free: last 15 days data, Premium: up to 365 days.
    return 365 if is_premium(context) else 15


async def reject_if_banned(update: Update, context: ContextTypes.DEFAULT_TYPE) -> bool:
    # When triggered via callback buttons, `update.message` can be None.
    msg = update.message
    if msg is None and update.callback_query is not None:
        msg = update.callback_query.message
    if msg is None:
        msg = update.effective_message

    if int(context.user_data.get("is_banned") or 0) == 1:
        if msg is not None:
            await msg.reply_text(t(current_lang(context), "banned"))
        return True
    return False


def get_session(context: ContextTypes.DEFAULT_TYPE) -> dict:
    session = context.user_data.get("session")
    return session if isinstance(session, dict) else {}


def set_session(context: ContextTypes.DEFAULT_TYPE, session: dict) -> None:
    context.user_data["session"] = session


def clear_session(context: ContextTypes.DEFAULT_TYPE) -> None:
    context.user_data.pop("session", None)


async def send_main_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    lang = current_lang(context)
    premium_banner = "\n💎 PREMIUM ACTIVE" if is_premium(context) else ""
    await show_screen(update, f"{t(lang, 'menu_welcome')}{premium_banner}\n\nChoose an option:", main_menu_keyboard(lang))


async def send_trade_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    lang = current_lang(context)
    premium_banner = "\n💎 PREMIUM ACTIVE" if is_premium(context) else ""
    await show_screen(update, f"{t(lang, 'menu_trade_title')}{premium_banner}", trade_menu_keyboard(lang))


async def send_stats_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    lang = current_lang(context)
    premium_banner = "\n💎 PREMIUM ACTIVE" if is_premium(context) else ""
    await show_screen(update, f"{t(lang, 'menu_stats_title')}{premium_banner}", stats_menu_keyboard(lang))


async def send_tools_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    lang = current_lang(context)
    premium_banner = "\n💎 PREMIUM ACTIVE" if is_premium(context) else ""
    await show_screen(update, f"{t(lang, 'menu_tools_title')}{premium_banner}", tools_menu_keyboard(lang))


async def send_account_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    lang = current_lang(context)
    premium_banner = "\n💎 PREMIUM ACTIVE" if is_premium(context) else ""
    await show_screen(update, f"{t(lang, 'menu_account_title')}{premium_banner}", account_menu_keyboard(lang))


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    if await reject_if_banned(update, context):
        return

    # Auto-assign admin to the first user who calls /start.
    if update.effective_user:
        await db_to_thread(context, ensure_admin, settings.db_path, update.effective_user.id)

    lang = context.user_data.get("lang")
    if not lang:
        # First-time onboarding: choose language.
        clear_session(context)
        set_session(context, {"mode": "onboarding_language"})
        await update.effective_message.reply_text(t("en", "select_language"), reply_markup=language_keyboard())
        return

    clear_session(context)
    await update.effective_message.reply_text(t(current_lang(context), "welcome"), reply_markup=main_menu_keyboard())


async def help_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    if await reject_if_banned(update, context):
        return
    clear_session(context)
    await show_screen(update, t(current_lang(context), "help_body"), section_back_home_keyboard())


async def language_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    if await reject_if_banned(update, context):
        return
    clear_session(context)
    set_session(context, {"mode": "language_change"})
    await show_screen(update, t(current_lang(context), "select_language"), language_keyboard())


def _ai_comment(roi_percent: Decimal) -> str:
    if roi_percent >= Decimal("7"):
        return "Strong trade"
    if roi_percent >= Decimal("0"):
        return "Good trade"
    return "Weak trade"


def _trade_prompt_trade_header(n_trades: int, trade_index: int) -> str:
    return f"📦 Trade {trade_index}/{n_trades}"


async def begin_new_trade(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    clear_session(context)
    set_session(
        context,
        {
            "mode": "new_trade",
            "step": "count_trades",
        },
    )
    if not is_premium(context):
        user_id = update.effective_user.id
        today_iso = date.today().isoformat()
        today_trades = await db_to_thread(context, count_trades_by_date, settings.db_path, user_id, today_iso)
        if today_trades >= 5:
            await show_screen(update, t(current_lang(context), "limit_free_daily_reached"), main_menu_keyboard(current_lang(context)))
            return

    await show_screen(update, t(current_lang(context), "trade_count_prompt"), trade_count_keyboard())


async def begin_calculator(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    clear_session(context)
    set_session(context, {"mode": "calc", "step": "buy_price"})
    await show_screen(update, f"⚡️ Arbitrage Calculator\n{t(current_lang(context), 'calc_buy_prompt')}", back_cancel_keyboard())


async def begin_escrow_analyzer(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    lang = current_lang(context)
    if not is_premium(context):
        clear_session(context)
        await show_screen(update, t(lang, "premium_feature_locked"), section_back_home_keyboard())
        return
    clear_session(context)
    session_payload = {"mode": "escrow", "step": "select_mode"}
    query = update.callback_query
    if query and query.message:
        session_payload["ui_chat_id"] = query.message.chat_id
        session_payload["ui_message_id"] = query.message.message_id
    set_session(context, session_payload)
    await show_session_screen(update, context, f"{t(lang, 'escrow_title')}\n\n{t(lang, 'escrow_select_mode')}", escrow_mode_keyboard(lang))


async def begin_daily_report(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    clear_session(context)
    user_id = update.effective_user.id
    today_iso = date.today().isoformat()

    trades = await db_to_thread(context, trades_by_date, settings.db_path, user_id, today_iso)
    if not trades:
        await show_screen(update, t(current_lang(context), "daily_none"), main_menu_keyboard(current_lang(context)))
        return

    # Compute weighted averages from stored qty.
    total_qty = sum(Decimal(str(t["qty"])) for t in trades)
    sum_buy_w = sum(Decimal(str(t["buy_price"])) * Decimal(str(t["qty"])) for t in trades)
    sum_sell_w = sum(Decimal(str(t["sell_price"])) * Decimal(str(t["qty"])) for t in trades)
    avg_buy = sum_buy_w / total_qty
    avg_sell = sum_sell_w / total_qty
    roi = compute_roi(avg_buy, avg_sell)
    profit_total = sum(Decimal(str(t["profit"])) for t in trades)

    msg = (
        f"{t(current_lang(context), 'daily_report_today')}\n"
        f"📦 Trades: {len(trades)}\n"
        f"📉 Avg Buy: ₹{fmt_money(avg_buy)}\n"
        f"📈 Avg Sell: ₹{fmt_money(avg_sell)}\n"
        f"💰 Profit: ₹{fmt_money(profit_total)}\n"
        f"📊 ROI: {fmt_money(roi)}%\n"
    )
    await show_screen(update, msg, main_menu_keyboard(current_lang(context)))

    # Surprise system (rare): triggers after daily report.
    try:
        chance = 0.2 if is_premium(context) else 0.05
        if random.random() < chance:
            surprise_key = "surprise_premium" if is_premium(context) else "surprise_free"
            await update.effective_message.reply_text(t(current_lang(context), surprise_key))
    except Exception:
        pass


async def begin_history(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    clear_session(context)
    user_id = update.effective_user.id

    start_date = (date.today() - timedelta(days=days_limit(context) - 1)).isoformat()
    trades = await db_to_thread(context, last_trades, settings.db_path, user_id, 10, start_date)
    if not trades:
        await show_screen(update, t(current_lang(context), "no_trades_yet"), main_menu_keyboard(current_lang(context)))
        return

    lines = [f"📜 History (last 10 | last {days_limit(context)} days)"]
    for i, t in enumerate(trades, start=1):
        profit = Decimal(str(t["profit"]))
        roi = Decimal(str(t["roi"]))
        qty = Decimal(str(t["qty"]))
        buy = Decimal(str(t["buy_price"]))
        sell = Decimal(str(t["sell_price"]))
        lines.append(
            f"{i}) {t['trade_date']} | Qty {fmt_money(qty, 4)} | Buy {fmt_money(buy)} | Sell {fmt_money(sell)} | "
            f"Profit ₹{fmt_money(profit)} ({fmt_money(roi)}%)"
        )

    await show_screen(update, "\n".join(lines), main_menu_keyboard(current_lang(context)))


async def begin_analytics(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    clear_session(context)
    user_id = update.effective_user.id

    start_date = (date.today() - timedelta(days=days_limit(context) - 1)).isoformat()
    summary = await db_to_thread(context, user_profit_summary_since, settings.db_path, user_id, start_date)
    if summary["trades"] == 0:
        await show_screen(update, t(current_lang(context), "analytics_none"), main_menu_keyboard(current_lang(context)))
        return

    msg = (
        f"📈 Analytics (Last {days_limit(context)} Days)\n"
        f"📦 Trades: {summary['trades']}\n"
        f"📦 Volume: {fmt_money(Decimal(str(summary['total_qty'])), 4)}\n"
        f"📉 Avg Buy: ₹{fmt_money(Decimal(str(summary['avg_buy'])))}\n"
        f"📈 Avg Sell: ₹{fmt_money(Decimal(str(summary['avg_sell'])))}\n"
        f"💰 Total Profit: ₹{fmt_money(Decimal(str(summary['total_profit'])))}\n"
        f"📊 ROI: {fmt_money(Decimal(str(summary['roi'])))}%\n"
    )
    await show_screen(update, msg, main_menu_keyboard(current_lang(context)))


def _format_trade_inputs_for_save(trades_inputs: list[dict]) -> list[TradeInputs] | None:
    parsed: list[TradeInputs] = []
    for t in trades_inputs:
        usdt_spent = t.get("usdt_spent")
        buy_price = t.get("buy_price")
        sell_price = t.get("sell_price")
        if usdt_spent is None or buy_price is None or sell_price is None:
            return None
        parsed.append(TradeInputs(usdt_spent=usdt_spent, buy_price=buy_price, sell_price=sell_price))
    return parsed


async def finalize_new_trade(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    user_id = update.effective_user.id

    session = get_session(context)
    trades_inputs = session.get("trades_inputs") or []
    parsed = _format_trade_inputs_for_save(trades_inputs)
    if not parsed:
        await update.effective_message.reply_text(t(current_lang(context), "missing_values"), reply_markup=main_menu_keyboard())
        clear_session(context)
        return

    today_iso = date.today().isoformat()

    # Enforce daily free trade limit again (safety against edge cases).
    if not is_premium(context):
        existing_today = await db_to_thread(
            context, count_trades_by_date, settings.db_path, user_id, today_iso
        )
        if existing_today + len(parsed) > 5:
            await update.effective_message.reply_text(
                t(current_lang(context), "limit_free_daily_reached"),
                reply_markup=main_menu_keyboard(),
            )
            clear_session(context)
            return

    # Compute totals.
    profit_total = sum((compute_trade_profit(t) for t in parsed), Decimal("0"))
    total_qty = sum((compute_qty(t) for t in parsed), Decimal("0"))
    weighted = compute_weighted_averages(parsed)
    avg_buy = weighted["avg_buy"]
    avg_sell = weighted["avg_sell"]
    roi_total = compute_roi(avg_buy, avg_sell)

    # Save each trade row in DB.
    for trade in parsed:
        qty = compute_qty(trade)
        await db_to_thread(
            context,
            add_trade,
            settings.db_path,
            user_id,
            today_iso,
            float(qty),
            float(trade.buy_price),
            float(trade.sell_price),
        )

    # AI-like response selection (no API):
    # profit -> positive messages, loss -> supportive, neutral -> normal.
    category = "neutral"
    if roi_total > 0:
        category = "profit"
    elif roi_total < 0:
        category = "loss"
    ai = translate_text(current_lang(context), category)
    clear_session(context)

    msg = (
        f"{t(current_lang(context), 'trade_saved_header')}\n\n"
        f"📉 Avg Buy: ₹{fmt_money(avg_buy)}\n"
        f"📈 Avg Sell: ₹{fmt_money(avg_sell)}\n"
        f"📦 Volume: {fmt_money(total_qty, 4)}\n"
        f"💰 Profit: ₹{fmt_money(profit_total)}\n"
        f"📊 ROI: {fmt_money(roi_total)}%\n\n"
        f"{t(current_lang(context), 'trade_saved_ai_label')} {ai}"
    )
    await update.effective_message.reply_text(msg, reply_markup=result_keyboard())

    # Surprise system (rare, no API): triggers after successful trade entry.
    try:
        chance = 0.2 if is_premium(context) else 0.05
        if random.random() < chance:
            surprise_key = "surprise_premium" if is_premium(context) else "surprise_free"
            await update.effective_message.reply_text(t(current_lang(context), surprise_key))
    except Exception:
        # Never break trade flow due to surprise errors.
        pass


async def nav_back(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    session = get_session(context)
    if not session:
        clear_session(context)
        await send_main_menu(update, context)
        return

    mode = session.get("mode")
    if mode == "new_trade":
        step = session.get("step")
        if step == "count_trades":
            clear_session(context)
            await send_main_menu(update, context)
            return

        n_trades = int(session["n_trades"])
        trade_index = int(session["trade_index"])  # 1-based

        # Going back rewinds the current trade fields so user can re-enter.
        if step == "trade_qty":
            if trade_index == 1:
                # Rewind to trade-count question.
                set_session(context, {"mode": "new_trade", "step": "count_trades"})
                await show_screen(update, t(current_lang(context), "trade_count_prompt"), trade_count_keyboard())
                return
            # Back to previous trade's sell_price
            session["trade_index"] = trade_index - 1
            session["step"] = "trade_sell_price"
            session["trades_inputs"][trade_index - 2]["sell_price"] = None
            # Clear later trades (including current)
            for j in range(trade_index - 1, n_trades):
                session["trades_inputs"][j] = {"usdt_spent": None, "buy_price": None, "sell_price": None}
        elif step == "trade_buy_price":
            session["step"] = "trade_qty"
            session["trades_inputs"][trade_index - 1]["buy_price"] = None
            session["trades_inputs"][trade_index - 1]["sell_price"] = None
        elif step == "trade_sell_price":
            session["step"] = "trade_buy_price"
            session["trades_inputs"][trade_index - 1]["sell_price"] = None

        n_trades = int(session["n_trades"])
        trade_index = int(session["trade_index"])
        header = _trade_prompt_trade_header(n_trades, trade_index)

        if session["step"] == "trade_qty":
            await show_screen(update, f"{header}\n{t(current_lang(context), 'trade_qty_prompt')}", back_cancel_keyboard())
        elif session["step"] == "trade_buy_price":
            await show_screen(update, f"{header}\n{t(current_lang(context), 'trade_buy_prompt')}", back_cancel_keyboard())
        elif session["step"] == "trade_sell_price":
            await show_screen(update, f"{header}\n{t(current_lang(context), 'trade_sell_prompt')}", back_cancel_keyboard())
        return

    if mode == "calc":
        step = session.get("step")
        if step == "buy_price":
            clear_session(context)
            await send_main_menu(update, context)
            return
        if step == "sell_price":
            session["step"] = "buy_price"
            await show_screen(update, t(current_lang(context), "calc_buy_prompt"), back_cancel_keyboard())
            return
        if step == "usdt_spent":
            session["step"] = "sell_price"
            await show_screen(update, t(current_lang(context), "calc_sell_prompt"), back_cancel_keyboard())
            return
    if mode == "escrow":
        step = session.get("step")
        if step == "select_mode":
            clear_session(context)
            await send_tools_menu(update, context)
            return
        if step == "qty":
            clear_session(context)
            await send_tools_menu(update, context)
            return
        if step == "rate":
            session["step"] = "qty"
            await show_session_screen(update, context, f"{t(current_lang(context), 'escrow_title')}\n\n{t(current_lang(context), 'escrow_qty_prompt')}", back_cancel_keyboard())
            return
        if step == "fee":
            session["step"] = "rate"
            await show_session_screen(update, context, f"{t(current_lang(context), 'escrow_title')}\n\n{t(current_lang(context), 'escrow_rate_prompt')}", back_cancel_keyboard())
            return
    clear_session(context)
    await send_main_menu(update, context)


async def nav_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    clear_session(context)
    await send_main_menu(update, context)


async def handle_trade_count_input(update: Update, context: ContextTypes.DEFAULT_TYPE, value: int) -> None:
    if not clamp_range(value, 1, 15):
        await update.effective_message.reply_text(
            t(current_lang(context), "invalid_count"),
            reply_markup=trade_count_keyboard(),
        )
        return

    n_trades = value
    if not is_premium(context):
        settings = context.application.bot_data["settings"]
        user_id = update.effective_user.id
        today_iso = date.today().isoformat()
        today_trades = await db_to_thread(context, count_trades_by_date, settings.db_path, user_id, today_iso)
        if today_trades + n_trades > 5:
            await update.effective_message.reply_text(
                t(current_lang(context), "limit_free_overflow"),
                reply_markup=trade_count_keyboard(),
            )
            return

    set_session(
        context,
        {
            "mode": "new_trade",
            "step": "trade_qty",
            "n_trades": n_trades,
            "trade_index": 1,
            "trades_inputs": [{"usdt_spent": None, "buy_price": None, "sell_price": None} for _ in range(n_trades)],
        },
    )
    await update.effective_message.reply_text(
        f"{_trade_prompt_trade_header(n_trades, 1)}\n{t(current_lang(context), 'trade_qty_prompt')}",
        reply_markup=back_cancel_keyboard(),
    )


async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    session = get_session(context)
    if not session:
        await update.effective_message.reply_text(t(current_lang(context), "type_command"), reply_markup=main_menu_keyboard())
        return

    user_text = (update.effective_message.text or "").strip()

    if session.get("mode") == "new_trade":
        step = session.get("step")
        n_trades = int(session.get("n_trades") or 0)
        trade_index = int(session.get("trade_index") or 0)
        trades_inputs = session.get("trades_inputs") or []

        if step == "count_trades":
            d = parse_decimal(user_text)
            if d is None:
                await update.effective_message.reply_text(t(current_lang(context), "invalid_number"), reply_markup=trade_count_keyboard())
                return
            # Accept integer only for counts.
            if d != d.to_integral_value():
                await update.effective_message.reply_text(t(current_lang(context), "invalid_count"), reply_markup=trade_count_keyboard())
                return
            await handle_trade_count_input(update, context, int(d))
            return

        d = parse_decimal(user_text)
        if d is None:
            await update.effective_message.reply_text(t(current_lang(context), "invalid_number"), reply_markup=back_cancel_keyboard())
            return
        if d <= 0:
            await update.effective_message.reply_text(t(current_lang(context), "value_gt_zero"), reply_markup=back_cancel_keyboard())
            return

        if step == "trade_qty":
            session["trades_inputs"][trade_index - 1]["usdt_spent"] = d
            session["step"] = "trade_buy_price"
            await update.effective_message.reply_text(
                f"{_trade_prompt_trade_header(n_trades, trade_index)}\n{t(current_lang(context), 'trade_buy_prompt')}",
                reply_markup=back_cancel_keyboard(),
            )
            return

        if step == "trade_buy_price":
            session["trades_inputs"][trade_index - 1]["buy_price"] = d
            session["step"] = "trade_sell_price"
            await update.effective_message.reply_text(
                f"{_trade_prompt_trade_header(n_trades, trade_index)}\n{t(current_lang(context), 'trade_sell_prompt')}",
                reply_markup=back_cancel_keyboard(),
            )
            return

        if step == "trade_sell_price":
            session["trades_inputs"][trade_index - 1]["sell_price"] = d
            # Advance or finalize.
            if trade_index >= n_trades:
                await finalize_new_trade(update, context)
                return
            session["trade_index"] = trade_index + 1
            session["step"] = "trade_qty"
            await update.effective_message.reply_text(
                f"{_trade_prompt_trade_header(n_trades, trade_index + 1)}\n{t(current_lang(context), 'trade_qty_prompt')}",
                reply_markup=back_cancel_keyboard(),
            )
            return

    if session.get("mode") == "calc":
        step = session.get("step")
        d = parse_decimal(user_text)
        if d is None or d <= 0:
            await update.effective_message.reply_text(t(current_lang(context), "value_gt_zero"), reply_markup=back_cancel_keyboard())
            return

        if step == "buy_price":
            session["buy_price"] = d
            session["step"] = "sell_price"
            await update.effective_message.reply_text(t(current_lang(context), "calc_sell_prompt"), reply_markup=back_cancel_keyboard())
            return
        if step == "sell_price":
            session["sell_price"] = d
            session["step"] = "usdt_spent"
            await update.effective_message.reply_text(t(current_lang(context), "calc_usdt_prompt"), reply_markup=back_cancel_keyboard())
            return
        if step == "usdt_spent":
            session["usdt_spent"] = d
            buy = session["buy_price"]
            sell = session["sell_price"]
            usdt_spent = session["usdt_spent"]

            trade = TradeInputs(usdt_spent=usdt_spent, buy_price=buy, sell_price=sell)
            qty = compute_qty(trade)
            profit = compute_trade_profit(trade)
            avg_buy = buy
            avg_sell = sell
            roi = compute_roi(avg_buy, avg_sell)
            clear_session(context)

            msg = (
                f"{t(current_lang(context), 'calc_result_header')}\n\n"
                f"📦 Qty: {fmt_money(qty, 6)}\n"
                f"📉 Buy: ₹{fmt_money(buy)}\n"
                f"📈 Sell: ₹{fmt_money(sell)}\n"
                f"💰 Profit: ₹{fmt_money(profit)}\n"
                f"📊 ROI: {fmt_money(roi)}%\n"
            )
            await update.effective_message.reply_text(msg, reply_markup=main_menu_keyboard())
            return

    if session.get("mode") == "escrow":
        step = session.get("step")
        d = parse_decimal(user_text)
        if d is None or d <= 0:
            await show_session_screen(update, context, t(current_lang(context), "value_gt_zero"), back_cancel_keyboard())
            return

        if step == "qty":
            session["qty"] = d
            session["step"] = "rate"
            await show_session_screen(update, context, f"{t(current_lang(context), 'escrow_title')}\n\n{t(current_lang(context), 'escrow_rate_prompt')}", back_cancel_keyboard())
            return
        if step == "rate":
            session["rate"] = d
            session["step"] = "fee"
            await show_session_screen(update, context, f"{t(current_lang(context), 'escrow_title')}\n\n{t(current_lang(context), 'escrow_fee_prompt')}", back_cancel_keyboard())
            return
        if step == "fee":
            session["fee"] = d
            qty = Decimal(str(session["qty"]))
            rate = Decimal(str(session["rate"]))
            fee_usdt = Decimal(str(session["fee"]))
            if qty <= 0:
                await show_session_screen(update, context, t(current_lang(context), "value_gt_zero"), back_cancel_keyboard())
                return
            # Updated escrow logic:
            # fee input is in USDT -> convert to INR, then apply effective-rate formulas.
            total = qty * rate
            fee_inr = fee_usdt * rate
            selected_mode = str(session.get("escrow_mode") or "buyer")
            if selected_mode == "buyer":
                effective_rate = (total + fee_inr) / qty
                mode_label = t(current_lang(context), "escrow_buyer_pays")
            else:
                effective_rate = (total - fee_inr) / qty
                mode_label = t(current_lang(context), "escrow_seller_pays")
            lang = current_lang(context)
            await asyncio.sleep(0.2)
            msg = (
                f"{t(lang, 'escrow_title')}\n\n"
                f"{mode_label}\n\n"
                f"📦 Qty: {fmt_money(qty, 6)}\n"
                f"💰 Rate: ₹{fmt_money(rate)}\n\n"
                f"💸 Fee: ₹{fmt_money(fee_inr)}\n"
                f"━━━━━━━━━━━━━━\n\n"
                f"{t(lang, 'escrow_effective')}: ₹{fmt_money(effective_rate)}"
            )
            await show_session_screen(update, context, msg, escrow_result_keyboard(lang))
            clear_session(context)
            return

    await update.effective_message.reply_text(t(current_lang(context), "unexpected"), reply_markup=main_menu_keyboard())


async def graph_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    if await reject_if_banned(update, context):
        return
    clear_session(context)
    user_id = update.effective_user.id

    dlimit = days_limit(context)
    series = await db_to_thread(context, daily_profit_series, settings.db_path, user_id, dlimit)
    if not series:
        await update.effective_message.reply_text(t(current_lang(context), "graph_none"), reply_markup=main_menu_keyboard())
        return

    dates = [s.trade_date for s in series]
    profits = [s.profit for s in series]

    def render_plot():
        fig = plt.figure(figsize=(10, 4))
        colors = ["#2ecc71" if p >= 0 else "#e74c3c" for p in profits]
        plt.bar(dates, profits, color=colors)
        plt.xticks(rotation=45, ha="right")
        plt.title("Profit by Day (Last 30 Days)")
        plt.tight_layout()
        buf = BytesIO()
        fig.savefig(buf, format="png", bbox_inches="tight")
        plt.close(fig)
        buf.seek(0)
        return buf

    buf = await asyncio.to_thread(render_plot)
    await update.effective_message.reply_photo(
        photo=InputFile(buf, filename="profit_graph.png"),
        caption=f"📊 Profit Graph (Last {dlimit} Days)",
        reply_markup=main_menu_keyboard(),
    )


async def month_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    if await reject_if_banned(update, context):
        return
    clear_session(context)
    set_session(context, {"mode": "month_range", "step": "await_days"})
    await show_screen(update, t(current_lang(context), "range_prompt"), days_range_keyboard())


async def render_month_report(update: Update, context: ContextTypes.DEFAULT_TYPE, days: int) -> None:
    settings = context.application.bot_data["settings"]
    user_id = update.effective_user.id
    today = date.today()
    start_date = (today - timedelta(days=days - 1)).isoformat()

    trades = await db_to_thread(context, user_trades_range, settings.db_path, user_id, start_date)
    if not trades:
        await show_screen(update, t(current_lang(context), "month_none").format(days=days), main_menu_keyboard(current_lang(context)))
        return

    # Compute totals.
    total_qty = sum(Decimal(str(tr["qty"])) for tr in trades)
    profit_total = sum(Decimal(str(tr["profit"])) for tr in trades)
    sum_buy_w = sum(Decimal(str(tr["buy_price"])) * Decimal(str(tr["qty"])) for tr in trades)
    sum_sell_w = sum(Decimal(str(tr["sell_price"])) * Decimal(str(tr["qty"])) for tr in trades)

    if total_qty == 0:
        avg_buy = Decimal("0")
        avg_sell = Decimal("0")
        roi = Decimal("0")
    else:
        avg_buy = sum_buy_w / total_qty
        avg_sell = sum_sell_w / total_qty
        roi = compute_roi(avg_buy, avg_sell)

    best, worst = await db_to_thread(context, best_worst_day, settings.db_path, user_id, days)

    lines = [
        f"📆 Last {days} Days Report",
        f"📦 Trades: {len(trades)}",
        f"📦 Volume: {fmt_money(total_qty, 4)}",
        f"💰 Total Profit: ₹{fmt_money(profit_total)}",
        f"📊 ROI: {fmt_money(roi)}%",
    ]
    if best:
        lines.append(f"🔥 Best Day: {best.trade_date} | ₹{fmt_money(Decimal(str(best.profit)))}")
    if worst:
        lines.append(f"🥶 Worst Day: {worst.trade_date} | ₹{fmt_money(Decimal(str(worst.profit)))}")

    await show_screen(update, "\n".join(lines), main_menu_keyboard(current_lang(context)))

    # Surprise system (rare): triggers after custom range report.
    try:
        chance = 0.2 if is_premium(context) else 0.05
        if random.random() < chance:
            surprise_key = "surprise_premium" if is_premium(context) else "surprise_free"
            await update.effective_message.reply_text(t(current_lang(context), surprise_key))
    except Exception:
        pass


def _user_handle(username: str | None, user_id: int) -> str:
    if username:
        return f"@{username}"
    return f"@user{user_id}"


async def leaderboard_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    if await reject_if_banned(update, context):
        return

    user_id = update.effective_user.id
    lang = current_lang(context)
    flags = await db_to_thread(context, get_user_flags, settings.db_path, user_id)
    if int(flags.get("leaderboard_consent_asked") or 0) != 1:
        set_session(context, {"mode": "leaderboard_consent_gate"})
        await show_screen(update, f"{t(lang, 'consent_leaderboard_title')}\n{t(lang, 'consent_leaderboard_body')}", consent_leaderboard_keyboard())
        return

    top = await db_to_thread(context, leaderboard_totals, settings.db_path, None, 10, True)
    rank_info = await db_to_thread(context, leaderboard_rank, settings.db_path, user_id, None, False)

    if rank_info.get("trade_count", 0) <= 0:
        await show_screen(update, t(lang, "leaderboard_no_trades"), main_menu_keyboard(lang))
        return

    lines: list[str] = [t(lang, "leaderboard_title")]
    if top:
        for i, entry in enumerate(top, start=1):
            profit_s = fmt_money(Decimal(str(entry["total_profit"])))
            lines.append(f"{i}. {_user_handle(entry.get('username'), entry['user_id'])} — ₹{profit_s}")
    else:
        lines.append(t(lang, "leaderboard_no_trades"))

    rank_line = t(lang, "leaderboard_you_rank").format(
        rank=rank_info.get("rank") or 0,
        profit=fmt_money(Decimal(str(rank_info.get("total_profit", 0.0)))),
    )
    lines.append("")
    lines.append(rank_line)

    await show_screen(update, "\n".join(lines), main_menu_keyboard(lang))


async def weekly_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    if await reject_if_banned(update, context):
        return

    user_id = update.effective_user.id
    lang = current_lang(context)
    flags = await db_to_thread(context, get_user_flags, settings.db_path, user_id)
    if int(flags.get("leaderboard_consent_asked") or 0) != 1:
        set_session(context, {"mode": "leaderboard_consent_gate"})
        await show_screen(update, f"{t(lang, 'consent_leaderboard_title')}\n{t(lang, 'consent_leaderboard_body')}", consent_leaderboard_keyboard())
        return

    start_date = (date.today() - timedelta(days=6)).isoformat()
    top = await db_to_thread(context, leaderboard_totals, settings.db_path, start_date, 10, True)
    rank_info = await db_to_thread(context, leaderboard_rank, settings.db_path, user_id, start_date, False)

    if rank_info.get("trade_count", 0) <= 0:
        await show_screen(update, t(lang, "weekly_no_trades"), main_menu_keyboard(lang))
        return

    lines: list[str] = [t(lang, "weekly_title")]
    if top:
        for i, entry in enumerate(top, start=1):
            profit_s = fmt_money(Decimal(str(entry["total_profit"])))
            lines.append(f"{i}. {_user_handle(entry.get('username'), entry['user_id'])} — ₹{profit_s}")
    else:
        lines.append(t(lang, "weekly_no_trades"))

    rank_line = t(lang, "weekly_you_rank").format(
        rank=rank_info.get("rank") or 0,
        profit=fmt_money(Decimal(str(rank_info.get("total_profit", 0.0)))),
    )
    lines.append("")
    lines.append(rank_line)

    await show_screen(update, "\n".join(lines), main_menu_keyboard(lang))


async def leaderboard_settings_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    if await reject_if_banned(update, context):
        return

    flags = await db_to_thread(context, get_user_flags, settings.db_path, update.effective_user.id)
    if int(flags.get("is_premium") or 0) != 1:
        await show_screen(update, "🔒 Leaderboard pause is available for premium users only.", main_menu_keyboard(current_lang(context)))
        return

    paused = int(flags.get("leaderboard_paused") or 0) == 1
    status = "Paused" if paused else "Visible"
    kb = InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton("⏸️ Pause", callback_data="lb_pause"),
                InlineKeyboardButton("▶️ Resume", callback_data="lb_resume"),
            ],
            [
                InlineKeyboardButton("❌ Remove My Name", callback_data="lb_remove"),
            ],
            [InlineKeyboardButton("⬅️ Back", callback_data="menu_home")],
        ]
    )
    await show_screen(update, f"🏆 Leaderboard Settings\nCurrent: {status}", kb)


async def reset_leaderboard_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    user_id = update.effective_user.id
    if not _is_authorized_admin(user_id):
        await update.effective_message.reply_text("❌ You are not authorized", reply_markup=main_menu_keyboard())
        return
    await db_to_thread(context, reset_leaderboard_all, settings.db_path)
    await update.effective_message.reply_text("✅ Leaderboard reset completed.", reply_markup=main_menu_keyboard())


async def export_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    if await reject_if_banned(update, context):
        return
    clear_session(context)
    user_id = update.effective_user.id

    dlimit = days_limit(context)
    start_date = (date.today() - timedelta(days=dlimit - 1)).isoformat()
    trades = await db_to_thread(context, user_trades_range, settings.db_path, user_id, start_date)
    summary = await db_to_thread(context, user_profit_summary_since, settings.db_path, user_id, start_date)

    if not trades:
        await update.effective_message.reply_text(t(current_lang(context), "export_none"), reply_markup=main_menu_keyboard())
        return

    def build_excel():
        wb = Workbook()
        ws = wb.active
        ws.title = "Trades"
        headers = ["Trade Date", "Created At", "Qty", "Buy Price", "Sell Price", "Profit", "ROI%"]
        ws.append(headers)
        for t in reversed(trades):
            ws.append(
                [
                    t["trade_date"],
                    t["created_at"],
                    float(t["qty"]),
                    float(t["buy_price"]),
                    float(t["sell_price"]),
                    float(t["profit"]),
                    float(t["roi"]),
                ]
            )

        # Summary sheet
        ws2 = wb.create_sheet("Summary")
        ws2.append(["Trades", summary["trades"]])
        ws2.append(["Total Qty", float(summary["total_qty"])])
        ws2.append(["Total Profit", float(summary["total_profit"])])
        ws2.append(["Avg Buy", float(summary["avg_buy"])])
        ws2.append(["Avg Sell", float(summary["avg_sell"])])
        ws2.append(["ROI%", float(summary["roi"])])

        # Column width tweaks
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 14

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    buf = await asyncio.to_thread(build_excel)
    await update.effective_message.reply_document(
        document=InputFile(buf, filename=f"trade_export_{user_id}.xlsx"),
        caption="📤 Export ready (Excel)",
        reply_markup=main_menu_keyboard(),
    )


async def menu_premium(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    clear_session(context)
    lang = current_lang(context)
    if not is_premium(context):
        await show_screen(update, t(lang, "premium_body"), main_menu_keyboard(lang))
        return
    await show_screen(update, t(lang, "premium_summary_title"), premium_dashboard_keyboard(lang))


def _summary_insight_key(summary: dict) -> str:
    roi = Decimal(str(summary.get("roi", 0)))
    if roi >= Decimal("5"):
        return "insight_positive"
    if roi >= Decimal("0"):
        return "insight_average"
    return "insight_warning"


def _calc_win_rate_and_streak(trades: list[dict]) -> tuple[Decimal, int]:
    if not trades:
        return Decimal("0"), 0
    wins = 0
    day_set: set[str] = set()
    for trade in trades:
        profit = Decimal(str(trade.get("profit", 0)))
        if profit > 0:
            wins += 1
        if trade.get("trade_date"):
            day_set.add(str(trade["trade_date"]))
    win_rate = (Decimal(wins) / Decimal(len(trades)) * Decimal("100")) if trades else Decimal("0")
    streak = 0
    d = date.today()
    while d.isoformat() in day_set:
        streak += 1
        d -= timedelta(days=1)
    return win_rate, streak


async def premium_insights_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    lang = current_lang(context)
    if not is_premium(context):
        await show_screen(update, t(lang, "premium_body"), main_menu_keyboard(lang))
        return
    await show_screen(update, t(lang, "premium_insights_title"), premium_insights_keyboard(lang))


async def premium_reports_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    lang = current_lang(context)
    if not is_premium(context):
        await show_screen(update, t(lang, "premium_body"), main_menu_keyboard(lang))
        return
    await show_screen(update, t(lang, "premium_reports_title"), premium_reports_keyboard(lang))


async def premium_insight_view(update: Update, context: ContextTypes.DEFAULT_TYPE, kind: str) -> None:
    settings = context.application.bot_data["settings"]
    lang = current_lang(context)
    user_id = update.effective_user.id
    days = days_limit(context)
    start_date = (date.today() - timedelta(days=days - 1)).isoformat()
    trades = await db_to_thread(context, user_trades_range, settings.db_path, user_id, start_date)
    summary = await db_to_thread(context, user_profit_summary_since, settings.db_path, user_id, start_date)
    best, worst = await db_to_thread(context, best_worst_day, settings.db_path, user_id, days)
    win_rate, streak = _calc_win_rate_and_streak(trades)
    if summary.get("trades", 0) == 0:
        await show_screen(update, t(lang, "no_data"), premium_insights_keyboard(lang))
        return

    text = t(lang, "premium_insights_title")
    if kind == "best":
        best_profit = Decimal(str(best.profit)) if best else Decimal("0")
        best_date = best.trade_date if best else "-"
        text = f"{t(lang, 'best_day_title')}\n\n📅 {best_date}\n💰 ₹{fmt_money(best_profit)}"
    elif kind == "worst":
        worst_profit = Decimal(str(worst.profit)) if worst else Decimal("0")
        worst_date = worst.trade_date if worst else "-"
        text = f"{t(lang, 'worst_day_title')}\n\n📅 {worst_date}\n💰 ₹{fmt_money(worst_profit)}"
    elif kind == "winrate":
        text = f"{t(lang, 'win_rate_title')}\n\n🎯 {fmt_money(win_rate)}%"
    elif kind == "streak":
        text = f"{t(lang, 'streak_title')}\n\n🔥 {streak}"
    elif kind == "volume":
        text = f"{t(lang, 'total_volume_title')}\n\n📦 {fmt_money(Decimal(str(summary['total_qty'])), 4)}"
    await show_screen(update, text, premium_insights_keyboard(lang))


async def premium_report_view(update: Update, context: ContextTypes.DEFAULT_TYPE, kind: str) -> None:
    settings = context.application.bot_data["settings"]
    lang = current_lang(context)
    user_id = update.effective_user.id
    days = days_limit(context)
    start_date = (date.today() - timedelta(days=days - 1)).isoformat()
    trades = await db_to_thread(context, user_trades_range, settings.db_path, user_id, start_date)
    summary = await db_to_thread(context, user_profit_summary_since, settings.db_path, user_id, start_date)
    if summary.get("trades", 0) == 0:
        await show_screen(update, t(lang, "no_data"), premium_reports_keyboard(lang))
        return

    total_profit = Decimal(str(summary["total_profit"]))
    total_volume = Decimal(str(summary["total_qty"]))
    win_rate, _streak = _calc_win_rate_and_streak(trades)
    active_days = len({str(tr.get("trade_date")) for tr in trades if tr.get("trade_date")}) or 1
    avg_daily = total_profit / Decimal(active_days)
    insight_line = t(lang, _summary_insight_key(summary))

    if kind == "full":
        text = (
            f"{t(lang, 'premium_summary_title')}\n\n"
            f"📅 {date.today().isoformat()}\n"
            f"💰 {t(lang, 'total_profit_title')}: ₹{fmt_money(total_profit)}\n"
            f"📦 {t(lang, 'total_volume_title')}: {fmt_money(total_volume, 4)}\n\n"
            f"{t(lang, 'roi_title')}: {fmt_money(Decimal(str(summary['roi'])))}%\n"
            f"🎯 {t(lang, 'win_rate_title')}: {fmt_money(win_rate)}%\n\n"
            f"{insight_line}"
        )
    elif kind == "avg_daily":
        text = f"{t(lang, 'avg_daily_profit_title')}\n\n📈 ₹{fmt_money(avg_daily)}"
    elif kind == "total_profit":
        text = f"{t(lang, 'total_profit_title')}\n\n💰 ₹{fmt_money(total_profit)}"
    else:
        text = f"{t(lang, 'total_volume_title')}\n\n📦 {fmt_money(total_volume, 4)}"
    await show_screen(update, text, premium_reports_keyboard(lang))


async def handle_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    if not query or not update.effective_user:
        return
    await query.answer()
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    if await reject_if_banned(update, context):
        return

    data = (query.data or "").strip()

    # Decorative/no-op button to keep UI layout clean.
    if data == "noop":
        return

    # Admin panel callbacks
    if data.startswith("admin_"):
        if not _is_authorized_admin(update.effective_user.id):
            await query.edit_message_text("❌ You are not authorized", reply_markup=main_menu_keyboard(current_lang(context)))
            return
        if data == "admin_subscriptions":
            subs = await db_to_thread(context, list_subscriptions, settings.db_path)
            if not subs:
                await query.edit_message_text("📊 Active Subscriptions\nNo subscriptions yet.", reply_markup=main_menu_keyboard(current_lang(context)))
                return
            lines = ["📊 Active Subscriptions"]
            for s in subs[:50]:
                uname = _user_handle(s.get("username"), s["user_id"])
                status = _status_text(int(s.get("is_premium") or 0), s.get("expiry_date"), int(s.get("is_paused") or 0))
                days_left = _days_left_from_expiry(s.get("expiry_date"))
                if status == "Active":
                    lines.append(f"{uname} — {days_left if days_left is not None else 0} days left")
                elif status == "Paused":
                    lines.append(f"{uname} — Paused")
                else:
                    lines.append(f"{uname} — Expired")
            await query.edit_message_text("\n".join(lines), reply_markup=main_menu_keyboard(current_lang(context)))
            return
        if data == "admin_premiumlist":
            plist = await db_to_thread(context, list_active_premium_users, settings.db_path)
            if not plist:
                await query.edit_message_text("💎 Premium List\nNo active premium users.", reply_markup=main_menu_keyboard(current_lang(context)))
                return
            lines = ["💎 Premium List"]
            for p in plist[:50]:
                uname = _user_handle(p.get("username"), p["user_id"])
                dl = _days_left_from_expiry(p.get("expiry_date"))
                suffix = "Paused" if int(p.get("is_paused") or 0) == 1 else f"{dl if dl is not None else 0} days left"
                lines.append(f"{uname} — {suffix}")
            await query.edit_message_text("\n".join(lines), reply_markup=main_menu_keyboard(current_lang(context)))
            return
        if data == "admin_check_help":
            await query.edit_message_text("Use: /check user_id", reply_markup=main_menu_keyboard(current_lang(context)))
            return
        if data == "admin_extend_help":
            await query.edit_message_text("Use: /extend user_id days", reply_markup=main_menu_keyboard(current_lang(context)))
            return
        if data == "admin_pause_help":
            await query.edit_message_text("Use: /pause user_id", reply_markup=main_menu_keyboard(current_lang(context)))
            return
        if data == "admin_resume_help":
            await query.edit_message_text("Use: /resume user_id", reply_markup=main_menu_keyboard(current_lang(context)))
            return
        if data == "admin_cancel_help":
            await query.edit_message_text("Use: /cancel user_id", reply_markup=main_menu_keyboard(current_lang(context)))
            return

    # User leaderboard visibility controls (premium-only).
    if data in {"lb_pause", "lb_resume", "lb_remove"}:
        flags = await db_to_thread(context, get_user_flags, settings.db_path, update.effective_user.id)
        if int(flags.get("is_premium") or 0) != 1 and data in {"lb_pause", "lb_resume"}:
            await query.edit_message_text(
                "🔒 Leaderboard pause is available for premium users only.",
                reply_markup=main_menu_keyboard(current_lang(context)),
            )
            return
        if data == "lb_remove":
            await db_to_thread(context, set_user_show_on_leaderboard, settings.db_path, update.effective_user.id, False)
            await db_to_thread(context, set_user_leaderboard_paused, settings.db_path, update.effective_user.id, False)
            await query.edit_message_text(
                "✅ You were removed from leaderboard.\nUse /leaderboard_settings to join again anytime.",
                reply_markup=main_menu_keyboard(current_lang(context)),
            )
            return

        paused = data == "lb_pause"
        await db_to_thread(context, set_user_leaderboard_paused, settings.db_path, update.effective_user.id, paused)
        status = "Paused" if paused else "Visible"
        kb = InlineKeyboardMarkup(
            [
                [
                    InlineKeyboardButton("⏸️ Pause", callback_data="lb_pause"),
                    InlineKeyboardButton("▶️ Resume", callback_data="lb_resume"),
                ],
                [InlineKeyboardButton("❌ Remove My Name", callback_data="lb_remove")],
                [InlineKeyboardButton("⬅️ Back", callback_data="menu_home")],
            ]
        )
        await query.edit_message_text(
            f"🏆 Leaderboard Settings\nCurrent: {status}",
            reply_markup=kb,
        )
        return

    # Language selection callback
    if data.startswith("lang_"):
        lang = data.replace("lang_", "").strip()
        if lang not in SUPPORTED_LANGS:
            lang = "en"
        await db_to_thread(context, set_user_language, settings.db_path, update.effective_user.id, lang)
        context.user_data["lang"] = lang
        # Always refresh immediately to the new-language main menu.
        clear_session(context)
        premium_banner = "\n💎 PREMIUM ACTIVE" if is_premium(context) else ""
        await query.edit_message_text(
            text=f"{t(current_lang(context), 'menu_welcome')}{premium_banner}\n\nChoose an option:",
            reply_markup=main_menu_keyboard(current_lang(context)),
        )
        return

    if data == "menu_language":
        clear_session(context)
        set_session(context, {"mode": "language_change"})
        await query.edit_message_text(
            text=t(current_lang(context), "select_language"),
            reply_markup=language_keyboard(),
        )
        return

    # Leaderboard consent
    if data == "consent_show":
        session_mode = get_session(context).get("mode")
        await db_to_thread(context, set_user_show_on_leaderboard, settings.db_path, update.effective_user.id, True)
        if session_mode == "leaderboard_consent_gate":
            clear_session(context)
            await query.edit_message_text(
                text=t(current_lang(context), "consent_leaderboard_thanks_show"),
                reply_markup=main_menu_keyboard(current_lang(context)),
            )
            await leaderboard_command(update, context)
        else:
            clear_session(context)
            await query.edit_message_text(
                text=t(current_lang(context), "consent_leaderboard_thanks_show"),
                reply_markup=main_menu_keyboard(current_lang(context)),
            )
        return
    if data == "consent_private":
        session_mode = get_session(context).get("mode")
        await db_to_thread(context, set_user_show_on_leaderboard, settings.db_path, update.effective_user.id, False)
        clear_session(context)
        if session_mode == "leaderboard_consent_gate":
            await query.edit_message_text(
                text=t(current_lang(context), "consent_leaderboard_thanks_private"),
                reply_markup=main_menu_keyboard(current_lang(context)),
            )
        else:
            await query.edit_message_text(
                text=t(current_lang(context), "consent_leaderboard_thanks_private"),
                reply_markup=main_menu_keyboard(current_lang(context)),
            )
        return

    # Custom range selection for /month
    if data.startswith("range_") and data[6:].isdigit():
        session = get_session(context)
        if session.get("mode") != "month_range":
            await query.edit_message_text(t(current_lang(context), "type_command"), reply_markup=main_menu_keyboard())
            return
        days = int(data.replace("range_", "").strip())
        if not is_premium(context) and days > 15:
            await query.edit_message_text(t(current_lang(context), "range_free_max"), reply_markup=days_range_keyboard())
            return
        clear_session(context)
        await render_month_report(update, context, days)
        return

    if data == "clear_yes":
        await db_to_thread(context, clear_user_trades, settings.db_path, update.effective_user.id)
        clear_session(context)
        await query.edit_message_text(
            text=t(current_lang(context), "clear_data_done"),
            reply_markup=account_menu_keyboard(current_lang(context)),
        )
        return

    if data == "clear_no":
        clear_session(context)
        await query.edit_message_text(
            text=t(current_lang(context), "clear_data_cancelled"),
            reply_markup=account_menu_keyboard(current_lang(context)),
        )
        return

    # Navigation buttons that should work in every state.
    if data == "nav_cancel":
        await nav_cancel(update, context)
        return
    if data == "nav_back":
        await nav_back(update, context)
        return

    # Trade count quick buttons.
    if data.startswith("tc_") and data[3:].isdigit():
        n = int(data[3:])
        # Only accept if in new_trade count stage.
        session = get_session(context)
        if session.get("mode") != "new_trade" or session.get("step") != "count_trades":
            await query.edit_message_text("📊 Start with `📊 New Trade` first.", reply_markup=main_menu_keyboard())
            return
        await handle_trade_count_input(update, context, n)
        return

    # Menu actions
    if data == "menu_home":
        clear_session(context)
        await send_main_menu(update, context)
        return
    if data == "menu_trade":
        await send_trade_menu(update, context)
        return
    if data == "menu_stats":
        await send_stats_menu(update, context)
        return
    if data == "menu_tools":
        await send_tools_menu(update, context)
        return
    if data == "menu_account":
        await send_account_menu(update, context)
        return
    if data == "menu_clear_data":
        clear_session(context)
        clear_prompt = (
            f"{t(current_lang(context), 'clear_data_title')}\n\n"
            f"{t(current_lang(context), 'clear_data_body')}"
        )
        kb = InlineKeyboardMarkup(
            [
                [
                    InlineKeyboardButton(t(current_lang(context), "clear_data_yes"), callback_data="clear_yes"),
                    InlineKeyboardButton("⬅️ Back", callback_data="nav_back"),
                ],
                [InlineKeyboardButton(t(current_lang(context), "clear_data_no"), callback_data="clear_no")],
            ]
        )
        await query.edit_message_text(text=clear_prompt, reply_markup=kb)
        return
    if data == "menu_leaderboard_settings":
        await leaderboard_settings_command(update, context)
        return
    if data == "menu_new_trade":
        await begin_new_trade(update, context)
        return
    if data == "menu_analytics":
        await begin_analytics(update, context)
        return
    if data == "menu_history":
        await begin_history(update, context)
        return
    if data == "menu_graph":
        await graph_command(update, context)
        return
    if data == "menu_calculator":
        await begin_calculator(update, context)
        return
    if data == "menu_escrow":
        await begin_escrow_analyzer(update, context)
        return
    if data in {"escrow_mode_buyer", "escrow_mode_seller"}:
        session = get_session(context)
        if session.get("mode") != "escrow":
            await send_tools_menu(update, context)
            return
        session["escrow_mode"] = "buyer" if data == "escrow_mode_buyer" else "seller"
        session["step"] = "qty"
        await show_session_screen(
            update,
            context,
            f"{t(current_lang(context), 'escrow_title')}\n\n{t(current_lang(context), 'escrow_qty_prompt')}",
            back_cancel_keyboard(),
        )
        return
    if data == "menu_daily":
        await begin_daily_report(update, context)
        return
    if data == "menu_reports":
        await month_command(update, context)
        return
    if data == "menu_leaderboard":
        await leaderboard_command(update, context)
        return
    if data == "menu_weekly":
        await weekly_command(update, context)
        return
    if data == "menu_premium":
        await menu_premium(update, context)
        return
    if data == "premium_insights_menu":
        await premium_insights_menu(update, context)
        return
    if data == "premium_reports_menu":
        await premium_reports_menu(update, context)
        return
    if data == "premium_insight_best":
        await premium_insight_view(update, context, "best")
        return
    if data == "premium_insight_worst":
        await premium_insight_view(update, context, "worst")
        return
    if data == "premium_insight_winrate":
        await premium_insight_view(update, context, "winrate")
        return
    if data == "premium_insight_streak":
        await premium_insight_view(update, context, "streak")
        return
    if data == "premium_insight_volume":
        await premium_insight_view(update, context, "volume")
        return
    if data == "premium_report_full":
        await premium_report_view(update, context, "full")
        return
    if data == "premium_report_avg_daily":
        await premium_report_view(update, context, "avg_daily")
        return
    if data == "premium_report_total_profit":
        await premium_report_view(update, context, "total_profit")
        return
    if data == "premium_report_total_volume":
        await premium_report_view(update, context, "total_volume")
        return
    if data == "menu_help":
        await help_cmd(update, context)
        return
    if data == "menu_export":
        await export_command(update, context)
        return


async def broadcast_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    if await reject_if_banned(update, context):
        return
    user_id = update.effective_user.id
    # If admin isn't set yet, make the first caller the admin.
    await db_to_thread(context, ensure_admin, settings.db_path, user_id)
    if not _is_authorized_admin(user_id):
        await update.effective_message.reply_text("❌ You are not authorized", reply_markup=main_menu_keyboard())
        return

    # Accept: /broadcast message...
    args = context.args or []
    if not args:
        clear_session(context)
        set_session(context, {"mode": "broadcast", "step": "await_text"})
        await update.effective_message.reply_text(t(current_lang(context), "broadcast_prompt"), reply_markup=back_cancel_keyboard())
        return
    text = " ".join(args).strip()

    users = await db_to_thread(context, all_users, settings.db_path)
    sent = 0
    failed = 0
    for uid, _username in users:
        try:
            await context.bot.send_message(chat_id=uid, text=f"📣 Broadcast:\n{text}")
            sent += 1
        except Exception:
            failed += 1
        await asyncio.sleep(0.02)

    await update.effective_message.reply_text(
        t(current_lang(context), "broadcast_done").format(sent=sent, failed=failed),
        reply_markup=main_menu_keyboard(),
    )


async def stats_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    if await reject_if_banned(update, context):
        return
    user_id = update.effective_user.id
    # If admin isn't set yet, make the first caller the admin.
    await db_to_thread(context, ensure_admin, settings.db_path, user_id)
    if not _is_authorized_admin(user_id):
        await update.effective_message.reply_text("❌ You are not authorized", reply_markup=main_menu_keyboard())
        return

    users = await db_to_thread(context, count_users, settings.db_path)
    trades = await db_to_thread(context, count_trades, settings.db_path)
    profit = await db_to_thread(context, sum_profit, settings.db_path)
    last_time = await db_to_thread(context, last_trade_time, settings.db_path)
    await update.message.reply_text(
        "🤖 Bot Stats\n"
        f"👥 Users: {users}\n"
        f"🧾 Trades: {trades}\n"
        f"💰 Total Profit: ₹{fmt_money(Decimal(str(profit)))}\n"
        f"🕒 Last Trade: {last_time or 'N/A'}",
        reply_markup=main_menu_keyboard(),
    )


async def users_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    if await reject_if_banned(update, context):
        return

    user_id = update.effective_user.id
    await db_to_thread(context, ensure_admin, settings.db_path, user_id)
    if not _is_authorized_admin(user_id):
        await update.effective_message.reply_text("❌ You are not authorized", reply_markup=main_menu_keyboard())
        return

    users = await db_to_thread(context, list_users, settings.db_path)
    lang = current_lang(context)
    if not users:
        await update.effective_message.reply_text(f"{t(lang, 'users_header')}\nNo users yet.", reply_markup=main_menu_keyboard())
        return

    top = users[:20]
    lines: list[str] = []
    for u in top:
        last_cleared = u.get("last_data_cleared_at")
        last_cleared_s = last_cleared.split("T")[0] if isinstance(last_cleared, str) and last_cleared else "N/A"
        lines.append(
            f"- {u['user_id']} | {u['username'] or 'no-username'} | lang:{u['language'] or 'en'} | prem:{u['is_premium']} | banned:{u['is_banned']} | clears:{u.get('data_clear_count', 0)} | last:{last_cleared_s}"
        )

    await update.effective_message.reply_text(
        f"{t(lang, 'users_header')}\n" + "\n".join(lines),
        reply_markup=main_menu_keyboard(),
    )


async def ban_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    if await reject_if_banned(update, context):
        return

    user_id = update.effective_user.id
    await db_to_thread(context, ensure_admin, settings.db_path, user_id)
    if not _is_authorized_admin(user_id):
        await update.effective_message.reply_text("❌ You are not authorized", reply_markup=main_menu_keyboard())
        return

    args = context.args or []
    if len(args) < 1:
        await update.effective_message.reply_text(t(current_lang(context), "ban_usage"), reply_markup=main_menu_keyboard())
        return
    try:
        target_id = int(args[0])
    except ValueError:
        await update.effective_message.reply_text(t(current_lang(context), "ban_usage"), reply_markup=main_menu_keyboard())
        return

    await db_to_thread(context, set_user_ban, settings.db_path, target_id, True)
    await update.effective_message.reply_text(t(current_lang(context), "ban_done"), reply_markup=main_menu_keyboard())


async def unban_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    if await reject_if_banned(update, context):
        return

    user_id = update.effective_user.id
    await db_to_thread(context, ensure_admin, settings.db_path, user_id)
    if not _is_authorized_admin(user_id):
        await update.effective_message.reply_text("❌ You are not authorized", reply_markup=main_menu_keyboard())
        return

    args = context.args or []
    if len(args) < 1:
        await update.effective_message.reply_text(t(current_lang(context), "unban_usage"), reply_markup=main_menu_keyboard())
        return
    try:
        target_id = int(args[0])
    except ValueError:
        await update.effective_message.reply_text(t(current_lang(context), "unban_usage"), reply_markup=main_menu_keyboard())
        return

    await db_to_thread(context, set_user_ban, settings.db_path, target_id, False)
    await update.effective_message.reply_text(t(current_lang(context), "unban_done"), reply_markup=main_menu_keyboard())


async def premium_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    if await reject_if_banned(update, context):
        return

    user_id = update.effective_user.id
    await db_to_thread(context, ensure_admin, settings.db_path, user_id)
    if not _is_authorized_admin(user_id):
        await update.effective_message.reply_text("❌ You are not authorized", reply_markup=main_menu_keyboard())
        return

    args = context.args or []
    if len(args) < 2:
        await update.effective_message.reply_text(t(current_lang(context), "premium_usage"), reply_markup=main_menu_keyboard())
        return
    try:
        target_id = int(args[0])
        days = int(args[1])
    except ValueError:
        await update.effective_message.reply_text(t(current_lang(context), "premium_usage"), reply_markup=main_menu_keyboard())
        return
    if days <= 0:
        await update.effective_message.reply_text(t(current_lang(context), "premium_usage"), reply_markup=main_menu_keyboard())
        return

    await db_to_thread(context, set_user_premium_days, settings.db_path, target_id, days)
    # Admin animation message (avoid spam: one message + edit).
    activating_msg = await update.effective_message.reply_text("💎 Activating Premium...", reply_markup=main_menu_keyboard(current_lang(context)))

    # Build user-facing premium activation message in target user's language.
    target_flags = await db_to_thread(context, get_user_flags, settings.db_path, target_id)
    target_lang = fallback_lang(target_flags.get("language"))
    labels = _premium_labels(target_lang)

    start_dt = datetime.utcnow()
    expiry_raw = target_flags.get("expiry_date")
    expiry_dt = None
    if expiry_raw:
        try:
            expiry_dt = datetime.fromisoformat(str(expiry_raw))
        except ValueError:
            expiry_dt = None
    # Fallback to computed expiry if DB value unavailable.
    if expiry_dt is None:
        expiry_dt = start_dt + timedelta(days=days)

    start_s = _format_pretty_date(start_dt)
    expiry_s = _format_pretty_date(expiry_dt)

    user_msg = (
        f"{labels['title']}\n\n"
        f"💎 PREMIUM ACTIVE\n\n"
        f"{labels['start']}: {start_s}\n"
        f"{labels['exp']}: {expiry_s}\n\n"
        f"{labels['features']}:\n"
        f"- Unlimited trades/day\n"
        f"- Up to 365 days reports\n"
        f"- Full analytics & leaderboard controls"
    )

    try:
        await context.bot.send_chat_action(chat_id=target_id, action=ChatAction.TYPING)
        await asyncio.sleep(0.75)
        await context.bot.send_message(chat_id=target_id, text=user_msg, reply_markup=main_menu_keyboard(target_lang))
        # Immediately send refreshed premium-aware menu so user never stays on stale UI.
        await context.bot.send_message(
            chat_id=target_id,
            text=f"{t(target_lang, 'menu_welcome')}\n💎 PREMIUM ACTIVE\n\nChoose an option:",
            reply_markup=main_menu_keyboard(target_lang),
        )
    except Exception:
        # Still complete admin flow even if target chat cannot be reached.
        pass

    # Edit admin message to finish.
    try:
        await activating_msg.edit_text(t(current_lang(context), "premium_done"), reply_markup=main_menu_keyboard(current_lang(context)))
    except Exception:
        # Fallback: send a final message.
        await update.effective_message.reply_text(t(current_lang(context), "premium_done"), reply_markup=main_menu_keyboard(current_lang(context)))


async def subscriptions_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    user_id = update.effective_user.id
    if not _is_authorized_admin(user_id):
        await update.effective_message.reply_text("❌ You are not authorized", reply_markup=main_menu_keyboard())
        return

    subs = await db_to_thread(context, list_subscriptions, settings.db_path)
    if not subs:
        await update.effective_message.reply_text("📊 Active Subscriptions\nNo subscriptions yet.", reply_markup=main_menu_keyboard())
        return

    lines = ["📊 Active Subscriptions"]
    for s in subs[:50]:
        uname = _user_handle(s.get("username"), s["user_id"])
        status = _status_text(int(s.get("is_premium") or 0), s.get("expiry_date"), int(s.get("is_paused") or 0))
        days_left = _days_left_from_expiry(s.get("expiry_date"))
        if status == "Active":
            lines.append(f"{uname} — {days_left if days_left is not None else 0} days left")
        elif status == "Paused":
            lines.append(f"{uname} — Paused")
        else:
            lines.append(f"{uname} — Expired")

    await update.effective_message.reply_text("\n".join(lines), reply_markup=main_menu_keyboard())


async def check_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    user_id = update.effective_user.id
    if not _is_authorized_admin(user_id):
        await update.effective_message.reply_text("❌ You are not authorized", reply_markup=main_menu_keyboard())
        return

    args = context.args or []
    if len(args) < 1:
        await update.effective_message.reply_text("❗ Use: /check user_id", reply_markup=main_menu_keyboard())
        return
    try:
        target_id = int(args[0])
    except ValueError:
        await update.effective_message.reply_text("❗ Invalid user_id", reply_markup=main_menu_keyboard())
        return

    rec = await db_to_thread(context, get_user_record, settings.db_path, target_id)
    if not rec:
        await update.effective_message.reply_text("❗ User not found", reply_markup=main_menu_keyboard())
        return

    uname = _user_handle(rec.get("username"), target_id)
    status = _status_text(int(rec.get("is_premium") or 0), rec.get("expiry_date"), int(rec.get("is_paused") or 0))
    expiry_raw = rec.get("expiry_date")
    try:
        expiry_dt = datetime.fromisoformat(str(expiry_raw)) if expiry_raw else None
    except ValueError:
        expiry_dt = None
    expiry_s = _format_pretty_date(expiry_dt)
    days_left = _days_left_from_expiry(expiry_raw)
    await update.effective_message.reply_text(
        f"👤 {uname}\n\n"
        f"💎 Premium: {status}\n"
        f"📅 Expiry: {expiry_s}\n"
        f"⏳ Days Left: {days_left if days_left is not None else 0}",
        reply_markup=main_menu_keyboard(),
    )


async def extend_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    user_id = update.effective_user.id
    if not _is_authorized_admin(user_id):
        await update.effective_message.reply_text("❌ You are not authorized", reply_markup=main_menu_keyboard())
        return
    args = context.args or []
    if len(args) < 2:
        await update.effective_message.reply_text("❗ Use: /extend user_id days", reply_markup=main_menu_keyboard())
        return
    try:
        target_id = int(args[0]); days = int(args[1])
    except ValueError:
        await update.effective_message.reply_text("❗ Invalid input", reply_markup=main_menu_keyboard())
        return
    if days <= 0:
        await update.effective_message.reply_text("❗ Days must be > 0", reply_markup=main_menu_keyboard())
        return
    await db_to_thread(context, extend_user_premium_days, settings.db_path, target_id, days)
    await update.effective_message.reply_text("✅ Subscription extended", reply_markup=main_menu_keyboard())


async def pause_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    user_id = update.effective_user.id
    if not _is_authorized_admin(user_id):
        await update.effective_message.reply_text("❌ You are not authorized", reply_markup=main_menu_keyboard())
        return
    args = context.args or []
    if len(args) < 1:
        await update.effective_message.reply_text("❗ Use: /pause user_id", reply_markup=main_menu_keyboard())
        return
    try:
        target_id = int(args[0])
    except ValueError:
        await update.effective_message.reply_text("❗ Invalid user_id", reply_markup=main_menu_keyboard())
        return
    await db_to_thread(context, pause_user_premium, settings.db_path, target_id)
    await update.effective_message.reply_text("⏸️ Subscription paused", reply_markup=main_menu_keyboard())


async def resume_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    user_id = update.effective_user.id
    if not _is_authorized_admin(user_id):
        await update.effective_message.reply_text("❌ You are not authorized", reply_markup=main_menu_keyboard())
        return
    args = context.args or []
    if len(args) < 1:
        await update.effective_message.reply_text("❗ Use: /resume user_id", reply_markup=main_menu_keyboard())
        return
    try:
        target_id = int(args[0])
    except ValueError:
        await update.effective_message.reply_text("❗ Invalid user_id", reply_markup=main_menu_keyboard())
        return
    await db_to_thread(context, resume_user_premium, settings.db_path, target_id)
    await update.effective_message.reply_text("▶️ Subscription resumed", reply_markup=main_menu_keyboard())


async def cancel_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    user_id = update.effective_user.id
    if not _is_authorized_admin(user_id):
        await update.effective_message.reply_text("❌ You are not authorized", reply_markup=main_menu_keyboard())
        return
    args = context.args or []
    if len(args) < 1:
        await update.effective_message.reply_text("❗ Use: /cancel user_id", reply_markup=main_menu_keyboard())
        return
    try:
        target_id = int(args[0])
    except ValueError:
        await update.effective_message.reply_text("❗ Invalid user_id", reply_markup=main_menu_keyboard())
        return
    await db_to_thread(context, cancel_user_premium, settings.db_path, target_id)
    await update.effective_message.reply_text("🛑 Subscription cancelled", reply_markup=main_menu_keyboard())


async def premiumlist_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    user_id = update.effective_user.id
    if not _is_authorized_admin(user_id):
        await update.effective_message.reply_text("❌ You are not authorized", reply_markup=main_menu_keyboard())
        return
    plist = await db_to_thread(context, list_active_premium_users, settings.db_path)
    if not plist:
        await update.effective_message.reply_text("💎 Premium List\nNo active premium users.", reply_markup=main_menu_keyboard())
        return
    lines = ["💎 Premium List"]
    for p in plist[:50]:
        uname = _user_handle(p.get("username"), p["user_id"])
        dl = _days_left_from_expiry(p.get("expiry_date"))
        suffix = "Paused" if int(p.get("is_paused") or 0) == 1 else f"{dl if dl is not None else 0} days left"
        lines.append(f"{uname} — {suffix}")
    await update.effective_message.reply_text("\n".join(lines), reply_markup=main_menu_keyboard())


async def global_premium_on_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    user_id = update.effective_user.id
    if not _is_authorized_admin(user_id):
        await update.effective_message.reply_text("❌ You are not authorized", reply_markup=main_menu_keyboard())
        return

    await db_to_thread(context, set_global_premium, settings.db_path, True)
    context.application.bot_data["global_premium"] = 1

    # Optional broadcast: /global_premium_on broadcast
    if context.args and context.args[0].lower() == "broadcast":
        users = await db_to_thread(context, all_users, settings.db_path)
        sent = 0
        failed = 0
        for uid, _username in users:
            try:
                await context.bot.send_message(chat_id=uid, text="🎉 Global Premium is now ON for everyone!")
                sent += 1
            except Exception:
                failed += 1
            await asyncio.sleep(0.02)
        await update.effective_message.reply_text(
            f"✅ Global premium enabled.\n📣 Broadcast sent: {sent}\n❌ Failed: {failed}",
            reply_markup=main_menu_keyboard(),
        )
        return

    await update.effective_message.reply_text(
        "✅ Global premium enabled for all users.",
        reply_markup=main_menu_keyboard(),
    )


async def global_premium_off_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    user_id = update.effective_user.id
    if not _is_authorized_admin(user_id):
        await update.effective_message.reply_text("❌ You are not authorized", reply_markup=main_menu_keyboard())
        return

    await db_to_thread(context, set_global_premium, settings.db_path, False)
    context.application.bot_data["global_premium"] = 0
    await update.effective_message.reply_text(
        "✅ Global premium disabled. User subscriptions remain unchanged.",
        reply_markup=main_menu_keyboard(),
    )


async def global_premium_status_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    user_id = update.effective_user.id
    if not _is_authorized_admin(user_id):
        await update.effective_message.reply_text("❌ You are not authorized", reply_markup=main_menu_keyboard())
        return

    enabled = await db_to_thread(context, get_global_premium, settings.db_path)
    status_text = "🟢 ON" if int(enabled or 0) == 1 else "🔴 OFF"
    await update.effective_message.reply_text(
        f"🌐 Global Premium Status: {status_text}",
        reply_markup=main_menu_keyboard(),
    )


async def admin_panel_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    user_id = update.effective_user.id
    if not _is_authorized_admin(user_id):
        await update.effective_message.reply_text("❌ You are not authorized", reply_markup=main_menu_keyboard())
        return
    admin_kb = InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton("📊 Subscriptions", callback_data="admin_subscriptions"),
                InlineKeyboardButton("💎 Premium List", callback_data="admin_premiumlist"),
            ],
            [
                InlineKeyboardButton("🔍 Check User", callback_data="admin_check_help"),
                InlineKeyboardButton("➕ Extend", callback_data="admin_extend_help"),
            ],
            [
                InlineKeyboardButton("⏸️ Pause", callback_data="admin_pause_help"),
                InlineKeyboardButton("▶️ Resume", callback_data="admin_resume_help"),
            ],
            [
                InlineKeyboardButton("🛑 Cancel", callback_data="admin_cancel_help"),
                InlineKeyboardButton("⬅️ Back", callback_data="menu_home"),
            ],
        ]
    )
    await update.effective_message.reply_text("👑 Admin Panel", reply_markup=admin_kb)


async def handle_broadcast_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    session = get_session(context)
    if session.get("mode") != "broadcast":
        return
    if session.get("step") != "await_text":
        return

    settings = context.application.bot_data["settings"]
    user_id = update.effective_user.id
    # Banned users cannot access broadcast flow.
    await ensure_user(update, context, settings)
    if await reject_if_banned(update, context):
        return
    if not _is_authorized_admin(user_id):
        clear_session(context)
        await update.effective_message.reply_text("❌ You are not authorized", reply_markup=main_menu_keyboard())
        return

    text = (update.effective_message.text or "").strip()
    if not text:
        await update.effective_message.reply_text(t(current_lang(context), "message_empty"), reply_markup=back_cancel_keyboard())
        return

    users = await db_to_thread(context, all_users, settings.db_path)
    sent = 0
    failed = 0
    for uid, _username in users:
        try:
            await context.bot.send_message(chat_id=uid, text=f"📣 Broadcast:\n{text}")
            sent += 1
        except Exception:
            failed += 1
        await asyncio.sleep(0.02)

    clear_session(context)
    await update.effective_message.reply_text(
        t(current_lang(context), "broadcast_done").format(sent=sent, failed=failed),
        reply_markup=main_menu_keyboard(),
    )


async def handle_text_router(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    # Broadcast flow is special: it only accepts one free-text message.
    settings = context.application.bot_data["settings"]
    await ensure_user(update, context, settings)
    if await reject_if_banned(update, context):
        return
    session = get_session(context)
    if session.get("mode") == "broadcast":
        await handle_broadcast_text(update, context)
        return
    await handle_text(update, context)


async def daily_motivation_job(context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    users = await db_to_thread(context, active_users_for_motivation, settings.db_path)
    for uid, lang in users:
        lang = fallback_lang(lang)
        message = t(lang, "motivation")
        try:
            await context.bot.send_message(chat_id=uid, text=message)
        except Exception:
            # Ignore individual send errors (user blocked bot etc).
            pass
        await asyncio.sleep(0.02)


async def _maybe_send_surprise(context: ContextTypes.DEFAULT_TYPE, user_id: int, lang: str, is_premium_flag: bool) -> None:
    """
    Rare surprise messages after trades/reports (free: 5%, premium: 20%).
    """
    chance = 0.2 if is_premium_flag else 0.05
    if random.random() >= chance:
        return
    try:
        surprise_key = "surprise_premium" if is_premium_flag else "surprise_free"
        await context.bot.send_message(chat_id=user_id, text=t(lang, surprise_key))
    except Exception:
        pass


async def daily_report_job(context: ContextTypes.DEFAULT_TYPE) -> None:
    settings = context.application.bot_data["settings"]
    today_iso = date.today().isoformat()
    users = await db_to_thread(context, active_users_for_motivation, settings.db_path)
    for uid, lang in users:
        lang = fallback_lang(lang)
        flags = await db_to_thread(context, get_user_flags, settings.db_path, int(uid))
        summary = await db_to_thread(context, user_profit_summary_since, settings.db_path, int(uid), today_iso)
        try:
            if summary["trades"] <= 0:
                msg = t(lang, "daily_report_job_none")
                await context.bot.send_message(chat_id=uid, text=msg)
            else:
                profit_s = fmt_money(Decimal(str(summary["total_profit"])))
                volume_s = fmt_money(Decimal(str(summary["total_qty"])), 4)
                roi_s = fmt_money(Decimal(str(summary["roi"])))
                body = t(lang, "daily_report_job_body").format(profit=profit_s, volume=volume_s, roi=roi_s)
                await context.bot.send_message(chat_id=uid, text=body)
                await _maybe_send_surprise(context, uid, lang, bool(flags.get("is_premium")))
        except Exception:
            pass
        await asyncio.sleep(0.02)


async def weekly_report_job(context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    Runs at 20:00 daily, sends weekly report only on Sundays (server-local).
    """
    # Monday=0 ... Sunday=6
    if datetime.now().weekday() != 6:
        return
    settings = context.application.bot_data["settings"]
    start_date = (date.today() - timedelta(days=6)).isoformat()
    users = await db_to_thread(context, active_users_for_motivation, settings.db_path)
    for uid, lang in users:
        lang = fallback_lang(lang)
        flags = await db_to_thread(context, get_user_flags, settings.db_path, int(uid))
        summary = await db_to_thread(context, user_profit_summary_since, settings.db_path, int(uid), start_date)
        try:
            if summary["trades"] <= 0:
                msg = t(lang, "weekly_report_job_none")
                await context.bot.send_message(chat_id=uid, text=msg)
            else:
                profit_s = fmt_money(Decimal(str(summary["total_profit"])))
                volume_s = fmt_money(Decimal(str(summary["total_qty"])), 4)
                avg_buy_s = fmt_money(Decimal(str(summary["avg_buy"])))
                avg_sell_s = fmt_money(Decimal(str(summary["avg_sell"])))
                roi_s = fmt_money(Decimal(str(summary["roi"])))
                best, worst = await db_to_thread(context, best_worst_day, settings.db_path, int(uid), 7)
                best_date = best.trade_date if best else "N/A"
                worst_date = worst.trade_date if worst else "N/A"
                best_profit_s = fmt_money(Decimal(str(best.profit)) if best else Decimal("0"))
                worst_profit_s = fmt_money(Decimal(str(worst.profit)) if worst else Decimal("0"))
                body = t(lang, "weekly_report_job_body").format(
                    profit=profit_s,
                    volume=volume_s,
                    avg_buy=avg_buy_s,
                    avg_sell=avg_sell_s,
                    roi=roi_s,
                    best_date=best_date,
                    best_profit=best_profit_s,
                    worst_date=worst_date,
                    worst_profit=worst_profit_s,
                )
                await context.bot.send_message(chat_id=uid, text=body)
                await _maybe_send_surprise(context, uid, lang, bool(flags.get("is_premium")))
        except Exception:
            pass
        await asyncio.sleep(0.02)


async def random_checkin_job(context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    Rare random check-in to keep users engaged.
    """
    settings = context.application.bot_data["settings"]
    users = await db_to_thread(context, active_users_for_motivation, settings.db_path)
    for uid, lang in users:
        if random.random() > 0.02:
            continue
        try:
            lang = fallback_lang(lang)
            await context.bot.send_message(chat_id=uid, text=t(lang, "checkin_message"))
        except Exception:
            pass
        await asyncio.sleep(0.02)


async def inactive_reminder_job(context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    Remind users who haven't logged a trade for 2 days (sent once when threshold is hit).
    """
    settings = context.application.bot_data["settings"]
    reminders = await db_to_thread(context, inactive_users_for_reminder, settings.db_path, 2)
    for uid, lang in reminders:
        try:
            lang = fallback_lang(lang)
            await context.bot.send_message(chat_id=uid, text=t(lang, "inactive_message"))
        except Exception:
            pass
        await asyncio.sleep(0.02)


def main() -> None:
    settings = load_settings()
    init_db(settings.db_path)
    global_premium = get_global_premium(settings.db_path)

    app = Application.builder().token(settings.bot_token).concurrent_updates(True).build()
    app.bot_data["settings"] = settings
    app.bot_data["global_premium"] = int(global_premium or 0)

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("help", help_cmd))
    app.add_handler(CommandHandler("language", language_cmd))
    app.add_handler(CommandHandler("graph", graph_command))
    app.add_handler(CommandHandler("month", month_command))
    app.add_handler(CommandHandler("leaderboard", leaderboard_command))
    app.add_handler(CommandHandler("weekly", weekly_command))
    app.add_handler(CommandHandler("leaderboard_settings", leaderboard_settings_command))
    app.add_handler(CommandHandler("reset_leaderboard", reset_leaderboard_command))
    app.add_handler(CommandHandler("export", export_command))
    app.add_handler(CommandHandler("broadcast", broadcast_command))
    app.add_handler(CommandHandler("stats", stats_command))
    app.add_handler(CommandHandler("users", users_command))
    app.add_handler(CommandHandler("ban", ban_command))
    app.add_handler(CommandHandler("unban", unban_command))
    app.add_handler(CommandHandler("premium", premium_command))
    app.add_handler(CommandHandler("subscriptions", subscriptions_command))
    app.add_handler(CommandHandler("check", check_command))
    app.add_handler(CommandHandler("extend", extend_command))
    app.add_handler(CommandHandler("pause", pause_command))
    app.add_handler(CommandHandler("resume", resume_command))
    app.add_handler(CommandHandler("cancel", cancel_command))
    app.add_handler(CommandHandler("premiumlist", premiumlist_command))
    app.add_handler(CommandHandler("global_premium_on", global_premium_on_command))
    app.add_handler(CommandHandler("global_premium_off", global_premium_off_command))
    app.add_handler(CommandHandler("global_premium_status", global_premium_status_command))
    app.add_handler(CommandHandler("admin", admin_panel_command))

    app.add_handler(CallbackQueryHandler(handle_callback))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text_router))

    # Daily motivation at 08:00.
    # Note: time is server-local.
    try:
        app.job_queue.run_daily(daily_motivation_job, time=dtime(hour=8, minute=0), name="daily_motivation")
        app.job_queue.run_daily(daily_report_job, time=dtime(hour=21, minute=0), name="daily_report")
        app.job_queue.run_daily(weekly_report_job, time=dtime(hour=20, minute=0), name="weekly_report")
        app.job_queue.run_daily(inactive_reminder_job, time=dtime(hour=8, minute=15), name="inactive_reminder")
        app.job_queue.run_daily(random_checkin_job, time=dtime(hour=13, minute=0), name="random_checkin")
    except Exception:
        pass

    app.run_polling(allowed_updates=Update.ALL_TYPES, poll_interval=0.5)


if __name__ == "__main__":
    main()
